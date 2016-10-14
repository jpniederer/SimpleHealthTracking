import openpyxl
import os

def test():
    return 5 * 5

def get_sheet():
    os.chdir('C:/Test/')
    workbook = openpyxl.load_workbook('HealthTracking.xlsx')
    return workbook.get_sheet_by_name('HealthForImport')

# Will iterate through the spreadsheet and create sleep objects.
def create_sleeps(sheet, lastRow):
    for i in range(6, lastRow):
        if sheet.cell(row=i, column=1).value == "Bed":
            dt = sheet.cell(row=i, column=2).value
            dt = dt.replace(hour=sheet.cell(row=i, column=3).value.hour)
            dt = dt.replace(minute=sheet.cell(row=i, column=3).value.minute)

# Will iterate through the spreadsheet and create checkins.
def create_checkins(sheet, lastRow):
    x = 5

def create_medicine_takens(sheet, lastRow):
    x = 6